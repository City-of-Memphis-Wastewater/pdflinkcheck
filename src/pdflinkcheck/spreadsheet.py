#!/usr/bin/env python3
# SPDX-License-Identifier: MIT

from __future__ import annotations
import re
from pathlib import Path
from typing import Dict, List, Optional, Union

from openpyxl import Workbook
from openpyxl.styles import Font

from .paths import PDFLINKCHECK_HOME
from pdflinkcheck.io import get_friendly_path, get_unique_human_time
from pdflinkcheck.helpers import PageRef

# ----------------- Helper Functions -----------------

def is_temp_pdf(pdf_name: str) -> bool:
    """Detect likely temp or unstable PDF filenames."""
    temp_patterns = [r'^tmp', r'^~', r'^[0-9a-fA-F]{8,}-', r'^temp']
    return any(re.match(pat, pdf_name, re.IGNORECASE) for pat in temp_patterns)

def sanitize_excel_text(text: str) -> str:
    """Remove characters illegal for Excel cells."""
    if not isinstance(text, str):
        text = str(text)
    return ''.join(c for c in text if c.isprintable() or c in '\t')

def convert_goto_link(link: Dict, pdf_path: str) -> str:
    """
    Convert internal GoTo link into a clickable file:// URI.
    Uses .as_uri() for robust cross-platform path encoding.
    """
    page_index = link.get('destination_page')
    base_uri = Path(pdf_path).absolute().as_uri()
    
    if page_index is None:
        return base_uri

    human_page = PageRef.from_index(int(page_index)).human
    return f"{base_uri}#page={human_page}"

def prepare_links_by_type(results: Dict, pdf_path: str = "") -> Dict[str, List[Dict]]:
    """Groups links and TOC items into categories."""
    payload = results.get("data", {}) if results else {}
    
    ext_links = payload.get('external_links', []) or []
    int_links = payload.get('internal_links', []) or []
    all_toc = payload.get('toc', []) or []

    grouped = {
        'Table of Contents': [],
        'Internal Links': [],
        'External Links': [],
        'Other': []
    }

    # 1. Process TOC
    for item in all_toc:
        raw_target = item.get('target_page')
        pg_human = PageRef.from_index(int(raw_target)).human if isinstance(raw_target, (int, float)) else "N/A"
        grouped['Table of Contents'].append({
            'col1': item.get('level', 1),
            'col2': sanitize_excel_text(item.get('title', 'Untitled')),
            'col3': pg_human
        })

    # 2. Process Internal Links
    for link in int_links:
        anchor = sanitize_excel_text(link.get('anchor_text', 'Link (No Text)'))
        raw_src = link.get('page')
        pg_src = PageRef.from_index(int(raw_src)).human if isinstance(raw_src, (int, float)) else "N/A"
        
        raw_dst = link.get('destination_page')
        pg_dst = PageRef.from_index(int(raw_dst)).human if isinstance(raw_dst, (int, float)) else "N/A"
        
        jump_url = convert_goto_link(link, pdf_path) if pdf_path else f"Page {pg_dst}"

        grouped['Internal Links'].append({
            'col1': pg_src, 
            'col2': pg_dst, 
            'col3': anchor, 
            'hyperlink': jump_url 
        })

    # 3. Process External Links
    for link in ext_links:
        anchor = sanitize_excel_text(link.get('anchor_text', 'Link (No Text)'))
        raw_src = link.get('page')
        pg_src = PageRef.from_index(int(raw_src)).human if isinstance(raw_src, (int, float)) else "N/A"
        
        url = link.get('url') or ''
        l_type = str(link.get('link_type', '')).lower()
        
        if "external" in l_type or "uri" in l_type:
            grouped['External Links'].append({
                'col1': pg_src, 'col2': anchor, 'col3': url, 'hyperlink': url
            })
        else:
            grouped['Other'].append({
                'col1': pg_src, 'col2': anchor, 'col3': url, 'hyperlink': url
            })

    return grouped

def _export_links_to_xlsx(grouped_links: Dict[str, List[Dict]], output_file: Path) -> bool:
    """Physical writer. Writes layout structural headers even if empty."""
    wb = Workbook()
    wb.properties.creator = "pdflinkcheck"
    wb.properties.title = "PDF Link Report"
    
    sheet_configs = [
        ('Table of Contents', ['Level', 'Title', 'Target Page']),
        ('Internal Links',   ['Source Page', 'Dest Page', 'Anchor Text', 'Jump Link']),
        ('External Links',   ['Source Page', 'Anchor Text', 'URL', 'External Link']),
        ('Other',            ['Source Page', 'Anchor Text', 'Reference', 'Link'])
    ]

    for sheet_name, headers in sheet_configs:
        rows = grouped_links.get(sheet_name, [])
        
        # Always construct the sheets so structural validations/tests find a physical artifact
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        
        # Bold headers
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Write rows cleanly inline
        for item in rows:
            row_data = [item['col1'], item['col2'], item['col3']]
            if 'hyperlink' in item:
                row_data.append(item['hyperlink'])
            
            ws.append(row_data)
            current_row = ws.max_row
            
            # Use dynamic column length safety assignment
            if 'hyperlink' in item and item['hyperlink'] != 'N/A':
                link_cell = ws.cell(row=current_row, column=len(row_data))
                link_cell.hyperlink = item['hyperlink']
                link_cell.style = 'Hyperlink'
            
            # String enforcement logic safely clamped to row range bounds
            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                val_str = str(cell.value or '')
                if val_str.startswith(('http', 'mhtml', 'mailto', 'file')):
                    cell.number_format = '@'

        # COLUMN AUTO-FIT
        sample_rows = list(ws.iter_rows(min_row=1, max_row=100, values_only=True))
        if sample_rows:
            for i, column_cells in enumerate(zip(*sample_rows)):
                column_letter = ws.cell(row=1, column=i+1).column_letter
                max_len = 0
                for val in column_cells:
                    if val:
                        max_len = max(max_len, len(str(val)))
                ws.column_dimensions[column_letter].width = min(max_len + 2, 70)

    # Clean default artifact sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    wb.save(output_file)
    print(f"XLSX exported successfully to {get_friendly_path(output_file)}")
    return True

def export_report_links_to_xlsx(results: Dict, pdf_path: Optional[Union[str, Path]], pdf_library_name: str, output_dir: Optional[Union[str, Path]] = None) -> Optional[Path]:
    """Main entry point for spreadsheet generation."""
    output_dir = Path(output_dir) if output_dir else PDFLINKCHECK_HOME
    output_dir.mkdir(parents=True, exist_ok=True)
    
    metadata = results.get("summary_metadata", {}) or {}
    file_ov = metadata.get("file_overview", {}) or {}
    
    pdf_path = (
        pdf_path or
        file_ov.get("source_path") or 
        file_ov.get("pdf_path") or 
        file_ov.get("processing_path") or 
        ""
    )
    
    grouped_links = prepare_links_by_type(results, pdf_path=pdf_path)
    
    pdf_name = file_ov.get("pdf_name") or file_ov.get("source_path") or "file"
    pdf_stem = Path(pdf_name).stem

    timestamp = get_unique_human_time()
    output_file = output_dir / f"{pdf_stem}_{pdf_library_name}_{timestamp}_report.xlsx"

    if _export_links_to_xlsx(grouped_links, output_file):
        return output_file
    return None